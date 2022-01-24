#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import openpyxl
from openpyxl.styles.alignment import Alignment
import os

# In[102]:


def addImage(path, imageCol):
    wb = openpyxl.load_workbook(filename = path)
    ws = wb.worksheets[0]

    #style number must be in first column to add image

    rows = range(2, ws.max_row+1)

    for row in rows:
        ws.row_dimensions[row].height = 150

    columns = range(1, ws.max_column+1)
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i in rows:
        try:
            jpegName = ws.cell(row = i, column = 1).value.strip()+"_front.jpg"
            image = os.path.join("N:/IMAGES/200xImages/",jpegName)

            img = openpyxl.drawing.image.Image(image)
            img.anchor = imageCol+ str(i)
            img.height = 200
            ws.add_image(img)
        except:
            pass
    wb.save(path)

